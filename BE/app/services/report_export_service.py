from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="0B3B60", end_color="0B3B60", fill_type="solid")
HEADER_FONT = Font(name="Times New Roman", bold=True, color="FFFFFF", size=13)
TITLE_FONT = Font(name="Times New Roman", bold=True, size=14)
BODY_FONT = Font(name="Times New Roman", size=13)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


class ReportExportService:

    def export_inventory_report(self, data: dict) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo tồn kho"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"BÁO CÁO TỒN KHO THUỐC - VT Y TẾ THÁNG {data['thang']}/{data['nam']}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = ["STT", "Tên thuốc/VTYT", "Đơn vị", "Tồn đầu kỳ", "Nhập trong kỳ", "Xuất trong kỳ", "Tồn cuối kỳ"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for i, item in enumerate(data["danh_sach"], 1):
            r = 3 + i
            ws.cell(row=r, column=1, value=i).border = THIN_BORDER
            ws.cell(row=r, column=1).alignment = CENTER_ALIGN
            ws.cell(row=r, column=2, value=item["ten_thuoc"]).border = THIN_BORDER
            ws.cell(row=r, column=3, value=item["don_vi"]).border = THIN_BORDER
            ws.cell(row=r, column=3).alignment = CENTER_ALIGN
            ws.cell(row=r, column=4, value=item["ton_dau_ky"]).border = THIN_BORDER
            ws.cell(row=r, column=4).alignment = CENTER_ALIGN
            ws.cell(row=r, column=5, value=item["nhap_trong_ky"]).border = THIN_BORDER
            ws.cell(row=r, column=5).alignment = CENTER_ALIGN
            ws.cell(row=r, column=6, value=item["xuat_trong_ky"]).border = THIN_BORDER
            ws.cell(row=r, column=6).alignment = CENTER_ALIGN
            ws.cell(row=r, column=7, value=item["ton_cuoi_ky"]).border = THIN_BORDER
            ws.cell(row=r, column=7).alignment = CENTER_ALIGN

        r_total = 3 + len(data["danh_sach"]) + 1
        ws.cell(row=r_total, column=1, value="TỔNG").font = Font(name="Times New Roman", bold=True, size=14)
        ws.cell(row=r_total, column=1).border = THIN_BORDER
        ws.cell(row=r_total, column=1).alignment = CENTER_ALIGN
        for col in range(2, 4):
            ws.cell(row=r_total, column=col).border = THIN_BORDER
        ws.cell(row=r_total, column=4, value=data["tong_ton_dau"]).border = THIN_BORDER
        ws.cell(row=r_total, column=4).alignment = CENTER_ALIGN
        ws.cell(row=r_total, column=4).font = Font(name="Times New Roman", bold=True, size=14)
        ws.cell(row=r_total, column=5, value=data["tong_nhap"]).border = THIN_BORDER
        ws.cell(row=r_total, column=5).alignment = CENTER_ALIGN
        ws.cell(row=r_total, column=5).font = Font(name="Times New Roman", bold=True, size=14)
        ws.cell(row=r_total, column=6, value=data["tong_xuat"]).border = THIN_BORDER
        ws.cell(row=r_total, column=6).alignment = CENTER_ALIGN
        ws.cell(row=r_total, column=6).font = Font(name="Times New Roman", bold=True, size=14)
        ws.cell(row=r_total, column=7, value=data["tong_ton_cuoi"]).border = THIN_BORDER
        ws.cell(row=r_total, column=7).alignment = CENTER_ALIGN
        ws.cell(row=r_total, column=7).font = Font(name="Times New Roman", bold=True, size=14)

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        for col in range(4, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        return wb
